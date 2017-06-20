#!/usr/bin/env python3

import sys
from astropy.io import fits as F
from openpyxl import load_workbook
import numpy as np

import argparse
parser = argparse.ArgumentParser()
parser.add_argument("tracefile", type=str, nargs='+')
parser.add_argument("-i", "--infile", type=str)
args = parser.parse_args()


CRIFORS = "/home/tom/work/crires/crifors.git/"

if args.infile:
    infile = args.infile
else:
    infile = CRIFORS+'data/cpmcfgWVLEN Table.xlsx'

print('Opening',infile)
wb = load_workbook(infile, data_only=True)
cfg = wb['cpmcfgWVLEN_Table.csv']
fitskeys= [c.value for c in cfg.rows[5]] # 6th row has FITS header names
setting_col = 2       # 3rd column is std setting name
X = np.arange(2048)

# open a second time with fomula intact
wb2 = load_workbook(infile)
cfg2 = wb2['cpmcfgWVLEN_Table.csv']

for tracefile in args.tracefile:
    print("Working on trace",tracefile)
    trace = F.open(tracefile)
    std = trace[0].header['HIERARCH ESO INS WLEN ID']
    print(std)
    assert(len(std.split('/'))==3)

    for i,r in enumerate(cfg.rows):
        if r[setting_col].value == std:
            print('found row %s'%i)
            break

    row = cfg2.rows[i]

    D = dict()
    for key,col in zip(fitskeys,row):
        if key and key.startswith('INS.WLEN.CENY'):
            D[key] = col

    for i in [1,2,3]:
        tdata = trace['CHIP%s'%i].data
        maxtracenb = max([foo[4] for foo in tdata])
        if maxtracenb > 9:
            #print([(t[4],np.polyval(t[0][::-1],X)[1024]) for t in tdata])
            tdata=tdata[1:]
        for alla, upper, lower, order, tracenb, wave in tdata:
            y = np.polyval(alla[::-1],X)[1024]
            if y and not np.isnan(y):
                D['INS.WLEN.CENY%s%s'%(i,maxtracenb-tracenb)].value = y

        
wb2.save('test.xlsx')
